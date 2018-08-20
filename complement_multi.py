import os
import openpyxl
import Google_complement
import multiprocessing
import time
import util.configure
import sys
import json
import warnings
warnings.filterwarnings('ignore')

path_reference = os.path.join(os.getcwd(), 'gene.xlsx')
reference = openpyxl.load_workbook(path_reference)
sheet = reference.active

path_result = os.path.join(os.getcwd(), 'result')

column = util.configure.column


def complement(lock, lower, upper, batch=10):
    print(lower, "~", upper)
    while lower < upper:
        list_result = []
        for i in range(lower, lower + batch + 1):
            if i > sheet.max_row:
                break
            name = sheet[column["expert"] + str(i)].value
            name = name if name is not None else ''
            affiliation = sheet[column["affiliation"] + str(i)].value
            affiliation = affiliation if affiliation is not None else ''
            search = (name + ' and ' + affiliation)
            print('-------------------------------------------------------')
            print(i, search)

            address = ''
            country = ''
            language = ''
            position = ''

            # # 搜索-补全email, phone
            email, phone = Google_complement.get_email_and_phone(search)

            # if len(affiliation) > 0:
            #     address = Google_complement.get_address(affiliation)
            #     country = Google_complement.get_country(affiliation)
            # else:
            #     pass

            # position = Google_complement.get_position(search)

            list_result.append((str(i), email, phone, address, country, language, position))
            print((str(i), email, phone, address, country, language, position))

        lock.acquire()
        try:
            with open(path_result, 'a', encoding='utf-8') as f:
                f.write(json.dumps(list_result))
        finally:
            lock.release()

        lower = min(upper, lower + batch + 1)


if __name__ == '__main__':
    begin_no = int(sys.argv[1])
    counts = int(sys.argv[2])
    num_of_processing = 4 if len(sys.argv) < 4 else int(sys.argv[3])
    have_done = 0

    if begin_no + counts > sheet.max_row:
        counts = sheet.max_row - begin_no + 1

    quarter = counts // num_of_processing

    lock = multiprocessing.Lock()

    arg_list = [
        (lock, have_done + begin_no, begin_no + quarter),
        (lock, have_done + begin_no + quarter + 1, begin_no + quarter * 2),
        (lock, have_done + begin_no + quarter * 2 + 1, begin_no + quarter * 3),
        (lock, have_done + begin_no + quarter * 3 + 1, begin_no + quarter * 4),
        (lock, have_done + begin_no + quarter * 4 + 1, begin_no + quarter * 5),
        (lock, have_done + begin_no + quarter * 5 + 1, begin_no + quarter * 6),
    ]

    for i in range(1, num_of_processing + 1):
        process = multiprocessing.Process(target=complement, args=arg_list[i - 1])
        process.start()
